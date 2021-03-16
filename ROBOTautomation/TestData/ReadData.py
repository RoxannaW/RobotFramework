#### Read testdata from excel to use in our testcase.

import openpyxl
vk = openpyxl.load_workbook("C://Users/rwijtsma/Documents/data.xlsx")

def fetch_number_of_rows(Sheetname):
    sh = vk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname, row, cell):
    sh = vk.[Sheetname]
    cell = sh.cell(row, cell)
    return cell.value

print(fetch_number_of_rows("Sheet1"))
print(fetch_cell_data("Sheet1",1,1))
